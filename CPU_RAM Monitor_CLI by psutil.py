import time
import psutil
import csv
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import argparse
import sys

# ==============================================================================
# 1. HELPER FUNCTIONS
# ==============================================================================

def get_pid():
    """
    ตรวจหา PID ของโปรเซสที่กำลังเทรน
    - ตรวจหาจากไฟล์ C:\\temp\\training_pid.txt สำหรับ MATLAB ก่อน
    - หากไม่เจอ จะค้นหาโปรเซส Python ที่กำลังรันไฟล์ .py
    """
    # --- ตรวจสอบ MATLAB ก่อน ---
    pid_file_path = "C:\\temp\\training_pid.txt"
    try:
        if os.path.exists(pid_file_path):
            with open(pid_file_path, "r") as f:
                pid = int(f.read().strip())
            proc = psutil.Process(pid)
            if proc.is_running() and "matlab" in proc.name().lower():
                return pid, f"MATLAB (PID: {pid}) CMD: {' '.join(proc.cmdline())}"
    except (FileNotFoundError, psutil.NoSuchProcess, ValueError, psutil.AccessDenied):
        pass # หากมีปัญหา ให้ข้ามไปหา Python

    # --- หากไม่เจอ MATLAB ให้ตรวจสอบ Python ---
    current_pid = psutil.Process().pid
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            if proc.pid == current_pid:
                continue

            name = proc.info['name'].lower()
            cmdline_list = proc.info.get('cmdline')
            cmdline = ' '.join(cmdline_list or []).lower()

            if ("python" in name or "python.exe" in name) and ".py" in cmdline:
                return proc.pid, f"Python: {' '.join(proc.info.get('cmdline') or [])}"
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    return None, None

def format_duration(seconds):
    """แปลงวินาทีเป็นรูปแบบ H:MM:SS.ms"""
    try:
        s_int = int(seconds)
        milliseconds = int((seconds - s_int) * 1000)
        hours, remainder = divmod(s_int, 3600)
        minutes, secs = divmod(remainder, 60)
        return f"{hours}:{minutes:02d}:{secs:02d}.{milliseconds:03d}"
    except (ValueError, TypeError):
        return str(seconds)

def get_update_interval(elapsed):
    """คำนวณช่วงเวลาการแสดงผลแบบ Buffered ตามเวลาที่ผ่านไป"""
    if elapsed <= 10: return 10
    elif elapsed <= 20: return 2
    elif elapsed <= 60: return 5
    elif elapsed <= 300: return 10
    elif elapsed <= 900: return 20
    else: return 30

def get_autosave_path(export_type, filename=None):
    """สร้างหรือสอบถาม path สำหรับ Auto-Save"""
    if not export_type:
        export_type = 'csv' # Default เป็น CSV
        
    if filename:
        full_filename = f"{filename}.{export_type}"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        full_filename = f"Data_{timestamp}.{export_type}"
        
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(downloads_path, exist_ok=True)
    return os.path.join(downloads_path, full_filename)

def auto_save_to_file(data, source, path):
    """บันทึกข้อมูล (Append) ลงในไฟล์ Excel หรือ CSV"""
    
    write_header = False
    if not os.path.exists(path) or (os.path.exists(path) and os.path.getsize(path) == 0):
        write_header = True

    all_data_to_save = data
    
    try:
        if path.lower().endswith('.xlsx'):
            if write_header:
                wb = Workbook()
                ws = wb.active
                # FIX: เปลี่ยนชื่อหัวตาราง
                ws.append(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
                wb.save(path)
                
            wb = load_workbook(path)
            ws = wb.active
            for row_data in all_data_to_save:
                formatted_row = [format_duration(row_data[0])] + list(row_data[1:])
                ws.append(formatted_row)
            wb.save(path)

        elif path.lower().endswith('.csv'):
            with open(path, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                if write_header:
                    # FIX: เปลี่ยนชื่อหัวตาราง
                    writer.writerow(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
                
                for row_data in all_data_to_save:
                    formatted_row = [format_duration(row_data[0])] + list(row_data[1:])
                    writer.writerow(formatted_row)
        
        return True
    except Exception as e:
        print(f"❌ Error saving data to {os.path.basename(path)}: {e}")
        return False

# ==============================================================================
# 2. CORE MONITORING LOGIC
# ==============================================================================

def monitor(samrate, display_mode, auto_save_path=None, total_elapsed_time=0.0):
    """
    ฟังก์ชันหลักสำหรับติดตามและบันทึกข้อมูล CPU/RAM
    
    :returns: (records, source, final_total_elapsed_time, final_auto_save_path)
    """
    print("🔍 Waiting for training process...")
    pid_file_path = "C:\\temp\\training_pid.txt"

    while True:
        pid, source = get_pid()
        if pid:
            break
        time.sleep(1)

    # ------------------------------------------------------------------
    # FIX: แยก Full Source และ Source สำหรับแสดงผลใน Terminal
    # ------------------------------------------------------------------
    full_source = source # นี่คือ Source เต็มรูปแบบที่จะใช้ในไฟล์ Excel/CSV
    
    # ตัด Source สำหรับ Terminal เพื่อให้ไม่ยาวเกินไป
    MAX_DISPLAY_LEN = 45 
    if len(full_source) > MAX_DISPLAY_LEN:
        display_source = full_source[:MAX_DISPLAY_LEN-3] + "..."
    else:
        display_source = full_source
        
    print(f"\n✅ Detected training from: {full_source}")
    # FIX: เปลี่ยนชื่อหัวตารางเป็น Time (H:MM:SS.ms)
    print(f"{'Time (H:MM:SS.ms)':<15} {'CPU (%)':<10} {'RAM (MB)':<12} {'Source':<45}") 
    # ------------------------------------------------------------------

    training_start = time.time()
    last_display_time = training_start
    data, buffer, samples = [], [], []
    is_matlab = "matlab" in source.lower()
    
    proc = psutil.Process(pid)
    
    # --- เริ่มต้นการนับ CPU Counter ---
    try:
        proc.cpu_percent(interval=None)
        time.sleep(0.1) 
    except (psutil.NoSuchProcess, psutil.AccessDenied) as e:
        print(f"❌ Cannot access initial CPU stats. Error: {e}")
        return [], full_source, 0.0, None 

    # --- กำหนดค่าสำหรับการคำนวณ Samplerate ---
    sample_interval = 0.1
    required_samples = max(1, int(samrate / sample_interval)) 
    samples_collected = 0

    while True:
        # --- เงื่อนไขการหยุด Monitor ---
        if is_matlab and not os.path.exists(pid_file_path):
            break
        if not psutil.pid_exists(pid):
            print("\nℹ️ Process PID not found. Stopping.")
            break

        # --- เก็บข้อมูล CPU/RAM ---
        try:
            start_of_sample = time.time()
            cpu = proc.cpu_percent(interval=None) / psutil.cpu_count()
            ram = proc.memory_info().rss / (1024 * 1024)
        except psutil.NoSuchProcess:
            break
        except Exception as e:
            break

        # --- ประมวลผลและแสดงข้อมูล ---
        current_session_elapsed = time.time() - training_start
        full_elapsed_seconds = total_elapsed_time + current_session_elapsed
        
        # NOTE: ใช้ full_source ในการบันทึก
        samples.append((full_elapsed_seconds, cpu, ram, full_source)) 
        samples_collected += 1
        
        # *** Auto-Save กลางทาง (ทุก 1 ชม. = 3600 วินาที) ***
        if current_session_elapsed >= 3600.0:
            
            # --- สร้างไฟล์อัตโนมัติถ้า auto_save_path เป็น None (คือเลือก 2) ---
            if auto_save_path is None:
                auto_save_path = get_autosave_path('csv') # สร้างไฟล์ CSV อัตโนมัติ
                print(f"\n🔔 Auto-save triggered! Auto-generating file: {os.path.basename(auto_save_path)}")

            if auto_save_path and len(samples) > 0:
                 # Flush samples และ buffer ก่อน auto-save
                # Note: เนื่องจาก samples ตอนนี้มี 4 คอลัมน์แล้ว (รวม full_source)
                if samples:
                    avg_cpu = sum(x[1] for x in samples) / len(samples) if samples else 0
                    avg_ram = sum(x[2] for x in samples) / len(samples) if samples else 0
                    timestamp = samples[-1][0]
                    # ใช้ full_source ในการสร้าง row สำหรับบันทึก/แสดงผล
                    row = (timestamp, avg_cpu, avg_ram, full_source) 
                    data.append(row)
                    samples.clear()
                    
                data.extend(buffer)
                buffer.clear()
                
                auto_save_to_file(data, full_source, auto_save_path)
                
                # รีเซ็ตค่าหลังจาก Auto-Save
                total_elapsed_time = full_elapsed_seconds
                training_start = time.time() # เริ่มนับเวลา session ใหม่
                data.clear()
                last_display_time = training_start
                print("🚨 Auto-Save completed. Monitoring session reset to continue tracking...\n")
                samples_collected = 0

        # บันทึก/แสดงผลตาม Sampling Rate
        if samples_collected >= required_samples:
            
            # คำนวณค่าเฉลี่ยของ samples ที่รวบรวมได้
            avg_cpu = sum(x[1] for x in samples) / len(samples) if samples else 0
            avg_ram = sum(x[2] for x in samples) / len(samples) if samples else 0
            current_full_elapsed = samples[-1][0] if samples else full_elapsed_seconds
            samples.clear()
            samples_collected = 0 # รีเซ็ตตัวนับ samples
            
            timestamp_str = format_duration(current_full_elapsed)
            
            # ใช้ full_source ในการสร้าง row สำหรับบันทึก/แสดงผล
            row = (current_full_elapsed, avg_cpu, avg_ram, full_source) 

            if display_mode == 1: # Real-time
                # FIX: ใช้ display_source สำหรับการแสดงผลใน Terminal
                print(f"{timestamp_str:<15} {avg_cpu:<10.2f} {avg_ram:<12.2f} {display_source:<45}") 
                data.append(row)
            else: # Buffered
                buffer.append(row)
                if time.time() - last_display_time >= get_update_interval(current_session_elapsed):
                    for b in buffer:
                        # FIX: ใช้ display_source สำหรับการแสดงผลใน Terminal
                        print(f"{format_duration(b[0]):<15} {b[1]:<10.2f} {b[2]:<12.2f} {display_source:<45}") 
                    data.extend(buffer)
                    buffer.clear()
                    last_display_time = time.time()
        
        # หน่วงเวลาที่เหลือ
        time_spent = time.time() - start_of_sample
        sleep_time = max(0, sample_interval - time_spent) # ใช้ sample_interval (0.1s) เป็นฐาน
        time.sleep(sleep_time)

    # Flush data ที่เหลือใน buffer
    if display_mode == 2 and buffer:
        for b in buffer:
            # FIX: ใช้ display_source สำหรับการแสดงผลใน Terminal
            print(f"{format_duration(b[0]):<15} {b[1]:<10.2f} {b[2]:<12.2f} {display_source:<45}") 
        data.extend(buffer)

    print("\n⏹️ Training stopped.")
    
    final_total_elapsed_time = total_elapsed_time + (time.time() - training_start)
    
    # คืนค่า auto_save_path ที่ถูกสร้างขึ้นอัตโนมัติกลับไปด้วย
    return data, full_source, final_total_elapsed_time, auto_save_path

# ==============================================================================
# 3. EXPORT FUNCTIONS (Non-Auto-Save)
# ==============================================================================

def export_excel(data, source, filename=None):
    """ส่งออกข้อมูลเป็นไฟล์ Excel (เขียนใหม่ทั้งหมด)"""
    if not filename:
        filename = input("Enter Excel filename (without extension): ").strip()
    if not filename:
        filename = f"monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    full_filename = f"{filename}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoring_Log"
    # FIX: เปลี่ยนชื่อหัวตาราง
    ws.append(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
    for row in data:
        formatted_row = [format_duration(row[0])] + list(row[1:])
        ws.append(formatted_row)
    ws.append([])
    ws.append(["Command/Source:", source])
    try:
        wb.save(full_filename)
        print(f"📁 Saved Excel to {os.path.abspath(full_filename)}")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")

def export_csv(data, source, filename=None):
    """ส่งออกข้อมูลเป็นไฟล์ CSV (เขียนใหม่ทั้งหมด)"""
    if not filename:
        filename = input("Enter CSV filename (without extension): ").strip()
    if not filename:
        filename = f"monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    full_filename = f"{filename}.csv"
    try:
        with open(full_filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            # FIX: เปลี่ยนชื่อหัวตาราง
            writer.writerow(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
            for row in data:
                formatted_row = [format_duration(row[0])] + list(row[1:])
                writer.writerow(formatted_row)
            writer.writerow([])
            writer.writerow(["Command/Source:", source])
        print(f"📁 Saved CSV to {os.path.abspath(full_filename)}")
    except Exception as e:
        print(f"❌ Error saving CSV file: {e}")


# ==============================================================================
# 4. MAIN INTERACTION LOGIC
# ==============================================================================

def main_cli(args):
    """ฟังก์ชันสำหรับโหมด CLI"""
    s = args.s
    mode = 1 if args.rt else 2

    auto_save_path = None
    if args.autosave:
        export_type = None
        if args.excel:
            export_type = 'xlsx'
        elif args.csv:
            export_type = 'csv'
        
        if not export_type:
            export_type = 'csv'
            
        auto_save_path = get_autosave_path(export_type, args.n)
        print(f"🛠️ Auto-Save mode enabled. Target file: {os.path.basename(auto_save_path)}")
        
    # รับค่า final_auto_save_path จาก monitor
    records, source, final_total_elapsed_time, auto_save_path = monitor(s, mode, auto_save_path=auto_save_path)

    # --- จัดการ Export (กรณีมีข้อมูลที่เหลือจากการ Auto-Save หรือเป็น Non-Auto-Save) ---
    if auto_save_path and (records or final_total_elapsed_time > 0.0):
        print(f"Saving final data to: {os.path.basename(auto_save_path)}")
        auto_save_to_file(records, source, auto_save_path)
    elif args.excel:
        export_excel(records, source, args.n)
    elif args.csv:
        export_csv(records, source, args.n)

    # --- จบการทำงานถ้ามี -end ---
    if args.end:
        print("👋 Exiting as requested by -end flag.")
        return
        
    # --- เมนูหลังจบการทำงาน (ถ้าไม่มี -end) ---
    while True:
        print("\n✅ Monitoring finished. What next?")
        print("1. Wait for new training")
        print("2. Export to Excel")
        print("3. Export to CSV")
        print("4. Restart from beginning")
        print("5. Exit")
        post = input("Choice: ").strip()
        if post == '1':
            print("\n" + "-"*40 + "\n")
            # เมื่อรอเทรนใหม่ ให้ส่ง auto_save_path เดิมไปเพื่อให้บันทึกต่อเนื่องได้
            records, source, final_total_elapsed_time, auto_save_path = monitor(s, mode, auto_save_path=auto_save_path, total_elapsed_time=0.0)
            continue
        elif post == '2':
            export_excel(records, source)
        elif post == '3':
            export_csv(records, source)
        elif post == '4':
            print("\n" + "="*40 + "\n")
            main_interactive()
            return
        elif post == '5':
            print("👋 Exiting...")
            return
        else:
            print("❌ Invalid choice.")


def main_interactive(prefilled_s=None):
    """ฟังก์ชันสำหรับโหมด Interactive (โต้ตอบกับผู้ใช้)"""
    s = 0.0
    if prefilled_s:
        s = prefilled_s
        print(f"⏱️ Sampling rate set to {s} sec via command line.")
    else:
        while True:
            try:
                s_input = input("⏱️ Set sampling rate (0.1–10.0) sec (recommended: 1.0): ")
                s = float(s_input)
                if 0.1 <= s <= 10.0: break
                else: print("❌ Invalid range. Try again.")
            except ValueError:
                print("❌ Invalid input. Try again.")

    while True: # Display mode loop
        print("\n📺 Select display mode:")
        print("1. Real-time display")
        print("2. Buffered display")
        print("3. Back to sampling rate")
        m = input("Choice: ").strip()

        if m == '3':
            main_interactive()
            return

        if m in ['1', '2']:
            mode = int(m)
            break
        else:
            print("❌ Invalid choice.")
            
    # --- Auto-Save Selection ---
    auto_save_path = None
    while True:
        print("\n💾 Select auto-save option (Saves data every 1 hour):")
        print("1. Select an Excel/CSV file to append data to")
        print("2. Do not select a file (Will auto-generate .csv file on the fly if monitoring runs long)")
        print("3. Back to display mode selection")
        a = input("Choice: ").strip()

        if a == '3':
            continue # Loop ไป Display mode

        if a == '1':
            while True:
                temp_path = input("Enter desired filename (e.g., mydata.xlsx or mydata.csv): ").strip()
                if not temp_path:
                    print("❌ Filename cannot be empty.")
                    continue
                
                if not (temp_path.lower().endswith('.xlsx') or temp_path.lower().endswith('.csv')):
                    print("❌ File must be .xlsx or .csv.")
                    continue
                
                export_type = 'xlsx' if temp_path.lower().endswith('.xlsx') else 'csv'
                filename_only = os.path.splitext(temp_path)[0]
                auto_save_path = get_autosave_path(export_type, filename=filename_only)
                print(f"✅ Auto-save file selected: {os.path.basename(auto_save_path)}")
                break
            break # ออกจาก Auto-Save loop
        
        if a == '2':
            print("✅ Auto-save file deferred. Will generate file if needed.")
            break # ออกจาก Auto-Save loop
        
        else:
            print("❌ Invalid choice.")
            
    # --- Select Action ---
    while True: # Action loop
        print("\n▶️ Select action:")
        print("1. Wait for training detection")
        print("2. Back to auto-save selection")
        action = input("Choice: ").strip()

        if action == '2':
            main_interactive(prefilled_s=s)
            return 

        if action == '1':
            # รับค่า final_auto_save_path จาก monitor
            records, source, final_total_elapsed_time, auto_save_path = monitor(s, mode, auto_save_path=auto_save_path)
            
            # จัดการบันทึกข้อมูลสุดท้าย
            # ใช้ auto_save_path ที่อาจถูกอัปเดตจาก monitor() แล้ว
            if auto_save_path and (records or final_total_elapsed_time > 0.0):
                print(f"Saving final data to: {os.path.basename(auto_save_path)}")
                auto_save_to_file(records, source, auto_save_path)
                
            # Post-monitoring loop
            while True: 
                print("\n✅ Monitoring finished. What next?")
                print("1. Wait for new training")
                print("2. Export to Excel")
                print("3. Export to CSV")
                print("4. Restart from beginning")
                print("5. Exit")
                post = input("Choice: ").strip()

                if post == '1':
                    print("\n" + "-"*40 + "\n")
                    # ส่ง auto_save_path ที่ถูกสร้างไปแล้ว
                    records, source, final_total_elapsed_time, auto_save_path = monitor(s, mode, auto_save_path=auto_save_path, total_elapsed_time=0.0) 
                    continue
                elif post == '2':
                    export_excel(records, source)
                elif post == '3':
                    export_csv(records, source)
                elif post == '4':
                    print("\n" + "="*40 + "\n")
                    main_interactive()
                    return 
                elif post == '5':
                    print("👋 Exiting...")
                    return
                else:
                    print("❌ Invalid choice.")
        else:
            print("❌ Invalid choice.")

# ==============================================================================
# 5. MAIN CLI ARGUMENT PARSER
# ==============================================================================

def main():
    """ฟังก์ชันหลักในการควบคุมโปรแกรมและจัดการ CLI arguments"""
    parser = argparse.ArgumentParser(
        description="Monitor training process CPU/RAM usage. Supports Auto-Save for long runs.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("-s", type=float, help="Set sampling rate (range: 0.1–10.0 sec).") 
    group_mode = parser.add_mutually_exclusive_group()
    group_mode.add_argument("-rt", action="store_true", help="Use real-time display mode.")
    group_mode.add_argument("-bf", action="store_true", help="Use buffered display mode.")
    
    # --- Auto-Save Flag ---
    parser.add_argument("-autosave", action="store_true", help="Enable automatic saving every 1 hour. Requires an export type.")
    
    group_export = parser.add_mutually_exclusive_group()
    group_export.add_argument("-excel", action="store_true", help="For non-autosave: Export to Excel after monitoring. \nFor autosave: Select Excel (.xlsx) file type.")
    group_export.add_argument("-csv", action="store_true", help="For non-autosave: Export to CSV after monitoring. \nFor autosave: Select CSV (.csv) file type.")
    
    parser.add_argument("-n", type=str, help="Filename for the export/autosave (without extension).")
    parser.add_argument("-end", action="store_true", help="End the program after monitoring and saving.")
    
    
    if len(sys.argv) == 1:
        main_interactive()
        return

    args, unknown = parser.parse_known_args()

    # --- CLI Validation ---
    if unknown:
        print(f"\n❌ Error: Unrecognized arguments: {' '.join(unknown)}")
        print("Here are the valid options:\n")
        parser.print_help()
        print("\n👋 Exiting.")
        return
        
    if args.autosave and not (args.excel or args.csv):
        print("\n❌ Error: The -autosave argument must be paired with an export type (-excel or -csv).")
        print("Here are the valid options:\n")
        parser.print_help()
        print("\n👋 Exiting.")
        return

    if args.n and not (args.excel or args.csv or args.autosave):
        print("\n❌ Error: The -n argument can only be used with an export flag (-excel, -csv) or -autosave.")
        print("Here are the valid options:\n")
        parser.print_help()
        print("\n👋 Exiting.")
        return

    if args.s is not None and (args.rt or args.bf):
        if not (0.1 <= args.s <= 10.0):
            print("\n❌ Error: Sampling rate (-s) must be between 0.1 and 10.0.")
            print("Here are the valid options:\n")
            parser.print_help()
            print("\n👋 Exiting.")
            return
        main_cli(args)
        return

    if args.s is not None and not any([args.rt, args.bf, args.excel, args.csv, args.n, args.end, args.autosave]):
        if not (0.1 <= args.s <= 10.0):
            print("\n❌ Error: Sampling rate (-s) must be between 0.1 and 10.0.")
            print("Here are the valid options:\n")
            parser.print_help()
            print("\n👋 Exiting.")
            return
        main_interactive(prefilled_s=args.s)
        return

    print("\n❌ Error: For CLI mode, both sampling rate (-s) and display mode (-rt or -bf) are required.")
    print("Here are the valid options:\n")
    parser.print_help()
    print("\n👋 Exiting.")


if __name__ == "__main__":
    main()