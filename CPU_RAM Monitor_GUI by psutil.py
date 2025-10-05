# -*- coding: utf-8 -*-
"""
แอป PyQt5 สำหรับมอนิเตอร์ CPU/RAM ของโปรเซส (MATLAB/Python)
- แสดงผลเป็นตาราง + กราฟ (matplotlib ฝังใน PyQt)
- เก็บข้อมูลเป็นช่วงเวลา (sampling rate ปรับได้)
- Auto-save ทุก 1 ชั่วโมง (หรือเมื่อถึงเงื่อนไข) ลงไฟล์ CSV/XLSX
    * ถ้า "ยังไม่ได้เลือกไฟล์" -> จะสร้างไฟล์ CSV อัตโนมัติในโฟลเดอร์ Downloads
- Final save ตอนจบ (append ต่อไฟล์เดิมถ้ามี autosave มาก่อน)
- ป้องกันกรณี "ไม่มีหัวตาราง" ด้วย _ensure_csv_header / _ensure_xlsx_header
"""

import sys
import psutil
import time
import threading
import csv
import os
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel,
    QFileDialog, QHBoxLayout, QDoubleSpinBox, QCheckBox,
    QTableWidget, QTableWidgetItem, QSplitter, QHeaderView
)
from PyQt5.QtCore import Qt, pyqtSignal, QObject

from openpyxl import Workbook, load_workbook

from matplotlib.backends.backend_qt5agg import (
    FigureCanvasQTAgg as FigureCanvas,
    NavigationToolbar2QT as NavigationToolbar
)
from matplotlib.figure import Figure


# ------------------------------
# ตัวกลางส่งสัญญาณ จาก thread ทำงานพื้นหลัง -> thread UI
# ------------------------------
class Worker(QObject):
    # ส่งข้อมูล batch ให้ UI อัปเดต พร้อม action (เช่น "flush")
    update_ui = pyqtSignal(list, str)
    # ส่งสัญญาณว่ามอนิเตอร์เสร็จสิ้น (เช่น โปรเซสตาย/จบ)
    finish_monitoring_signal = pyqtSignal(str)


# ------------------------------
# วิดเจ็ตพล็อตกราฟ CPU/RAM ด้วย matplotlib
# ------------------------------
class PlotCanvas(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Figure 1 อัน 2 แกน: CPU (บน) / RAM (ล่าง) และแชร์แกน X (เวลา)
        self.figure = Figure()
        self.ax_cpu, self.ax_ram = self.figure.subplots(2, 1, sharex=True)

        # Canvas + Toolbar ฝังใน PyQt
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)

        # Layout วาง toolbar เหนือกราฟ
        layout = QVBoxLayout()
        layout.addWidget(self.toolbar)
        layout.addWidget(self.canvas)
        self.setLayout(layout)

        # ตั้งค่าชื่อและแกน
        self.figure.suptitle("CPU and RAM Usage Over Time")
        self.ax_cpu.set_ylabel("CPU Usage (%)")
        self.ax_cpu.grid(True)
        self.ax_ram.set_ylabel("RAM Usage (MB)")
        self.ax_ram.set_xlabel("Time (H:MM:SS)")
        self.ax_ram.grid(True)
        self.figure.tight_layout(rect=[0, 0.03, 1, 0.95])

    def plot(self, timestamps, cpu_vals, ram_vals, is_real_time=True):
        # ล้างของเดิมก่อนพล็อต
        self.ax_cpu.clear()
        self.ax_ram.clear()

        # ถ้า real-time และข้อมูลยาวมาก ให้ตัดเหลือท้ายๆ เพื่อประสิทธิภาพการวาด
        if is_real_time and len(timestamps) > 1000:
            timestamps = timestamps[-1000:]
            cpu_vals = cpu_vals[-1000:]
            ram_vals = ram_vals[-1000:]

        # พล็อต 2 เส้น: CPU (%) / RAM (MB)
        self.ax_cpu.plot(timestamps, cpu_vals, '-', label='CPU (%)', color='tab:blue')
        self.ax_cpu.set_ylabel("CPU Usage (%)")
        self.ax_cpu.grid(True)

        self.ax_ram.plot(timestamps, ram_vals, '-', label='RAM (MB)', color='tab:orange')
        self.ax_ram.set_ylabel("RAM Usage (MB)")
        self.ax_ram.set_xlabel("Time (H:MM:SS)")
        self.ax_ram.grid(True)

        self.figure.suptitle("CPU and RAM Usage Over Time")
        self.figure.tight_layout(rect=[0, 0.03, 1, 0.95])
        self.canvas.draw()

    def reset_graph(self):
        # ล้างกราฟและตั้ง label ใหม่ (ใช้เวลาปิด plotting หรือ reset ตาราง)
        self.ax_cpu.clear()
        self.ax_ram.clear()
        self.ax_cpu.grid(True)
        self.ax_ram.grid(True)
        self.figure.suptitle("CPU and RAM Usage Over Time")
        self.ax_cpu.set_ylabel("CPU Usage (%)")
        self.ax_ram.set_ylabel("RAM Usage (MB)")
        self.ax_ram.set_xlabel("Time (H:MM:SS)")
        self.canvas.draw()


# ------------------------------
# วิดเจ็ตหลักของแอป
# ------------------------------
class MonitorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CPU/RAM Monitor by psutil")
        self.resize(1100, 700)

        # โฟลเดอร์ชั่วคราวสำหรับไฟล์ pid MATLAB (ถ้าใช้ร่วม)
        temp_dir = "C:\\temp"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        # ---------- สถานะหลัก ----------
        self.monitoring = False                 # กำลังมอนิเตอร์อยู่หรือไม่
        self.training_source = "Manual"         # แหล่งที่มา/คำสั่งแสดงในไฟล์ผลลัพธ์
        self.training_pid = None                # PID ของโปรเซสที่ติดตาม
        self.data = []                          # ข้อมูลที่แสดงแล้วในตาราง/กราฟ
        self.buffered_data = []                 # บัฟเฟอร์สะสมก่อน flush
        self.sampling_rate = 1.0                # คาบเวลาเก็บข้อมูล (วินาที)
        self.training_start_time = None         # เวลาเริ่มนับของ session ปัจจุบัน
        self.last_update_time = time.time()     # เวลา flush ล่าสุด
        self.update_interval = 2                # ช่วงเวลาระหว่างการ flush (โหมด buffered)
        self.initial_buffer_flushed = False     # เคย flush ครั้งแรกหรือยัง
        self.idle_start_time = None             # ใช้ขยายต่อได้ ถ้าต้อง detect idle
        self.IDLE_THRESHOLD_SECONDS = 30
        self.auto_save_path = None              # path ปลายทาง autosave/final save

        # เวลา cumulative ของทุก session (หลังจาก autosave จะ reset session time)
        self.total_elapsed_time = 0.0

        # ---------- ธงป้องกันเหตุ race/ซ้ำ ----------
        self._finish_emitted = False            # กันส่ง finish ซ้ำเมื่อโปรเซสจบ
        self._is_finalizing = False             # กัน reentry ใน finish_monitoring
        self._final_written = False             # กันเขียนไฟล์ไฟนอลซ้ำ
        self._autosave_written = False          # เคย autosave ระหว่างทางแล้วหรือยัง

        # Worker (สัญญาณระหว่าง thread)
        self.worker = Worker()
        self.worker.update_ui.connect(self.update_ui)
        self.worker.finish_monitoring_signal.connect(self.finish_monitoring)

        # ---------- ตารางแสดงผล ----------
        self.table = QTableWidget(0, 4)
        self.table.setWordWrap(False)
        self.table.setHorizontalHeaderLabels(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
        self.table.verticalHeader().setVisible(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)

        # ---------- Label สถานะ ----------
        self.status_label = QLabel("Status: Idle")
        self.source_label = QLabel("")

        # ---------- คอนโทรล UI ----------
        self.sampling_spinbox = QDoubleSpinBox()
        self.sampling_spinbox.setRange(0.1, 10.0)
        self.sampling_spinbox.setValue(1.0)
        self.sampling_spinbox.setSingleStep(0.1)

        self.auto_start_checkbox = QCheckBox("Start Detection Automatically")
        self.enable_plot_checkbox = QCheckBox("Enable Plotting")
        self.enable_plot_checkbox.setChecked(True)
        self.enable_plot_checkbox.stateChanged.connect(self.toggle_plot_options)

        self.plot_mode_checkbox = QCheckBox("Plot Graph After Training Ends")  # ถ้าเลือก -> ไม่พล็อตระหว่างทาง
        self.plot_mode_checkbox.setChecked(False)

        self.buffer_mode_checkbox = QCheckBox("Mode (tick=real-time, untick=buffered)")  # โหมด flush ทันที/เป็นช่วง
        self.buffer_mode_checkbox.setChecked(False)

        # ปุ่มต่างๆ
        self.btn_reset = QPushButton("Reset Table")
        self.btn_export_excel = QPushButton("Export to Excel")
        self.btn_export_csv = QPushButton("Export to CSV")
        self.btn_save_graph = QPushButton("Save Graph")
        self.btn_exit = QPushButton("Exit")

        self.btn_select_autosave = QPushButton("Select Auto-Save File")
        self.auto_save_file_label = QLabel("No file selected")
        self.btn_select_autosave.clicked.connect(self.select_autosave_file)

        # ผูกเหตุการณ์ปุ่มหลัก
        self.btn_reset.clicked.connect(self.reset_table)
        self.btn_export_excel.clicked.connect(self.export_excel)
        self.btn_export_csv.clicked.connect(self.export_csv)
        self.btn_save_graph.clicked.connect(self.save_graph)
        self.btn_exit.clicked.connect(self.close)

        # วิดเจ็ตกราฟ
        self.graph = PlotCanvas(self)

        # จัด Layout ทั้งหน้า
        self.setup_ui()
        self.toggle_plot_options()

        # สตาร์ท thread หลักสำหรับมอนิเตอร์ (background)
        threading.Thread(target=self.monitor_loop, daemon=True).start()

    # ------------------------------
    # helper: แปลงวินาที float -> "H:MM:SS.mmm" เพื่อแสดง/บันทึก
    # ------------------------------
    def format_duration(self, seconds):
        try:
            s_int = int(seconds)
            milliseconds = int((seconds - s_int) * 1000)
            hours, remainder = divmod(s_int, 3600)
            minutes, secs = divmod(remainder, 60)
            return f"{hours}:{minutes:02d}:{secs:02d}.{milliseconds:03d}"
        except (ValueError, TypeError):
            return str(seconds)

    # ------------------------------
    # จัดวางเลย์เอาต์ UI
    # ------------------------------
    def setup_ui(self):
        layout = QVBoxLayout()

        # แถวปุ่มควบคุมด้านล่าง
        control_layout = QHBoxLayout()
        control_layout.addWidget(QLabel("Sampling Rate (s):"))
        control_layout.addWidget(self.sampling_spinbox)
        control_layout.addStretch()
        control_layout.addWidget(self.btn_reset)
        control_layout.addWidget(self.btn_export_excel)
        control_layout.addWidget(self.btn_export_csv)
        control_layout.addWidget(self.btn_save_graph)
        control_layout.addWidget(self.btn_select_autosave)
        control_layout.addWidget(self.auto_save_file_label)
        control_layout.addWidget(self.btn_exit)

        # แถวเช็คบ็อกซ์ตัวเลือก
        checkbox_layout = QHBoxLayout()
        checkbox_layout.addWidget(self.auto_start_checkbox)
        checkbox_layout.addWidget(self.enable_plot_checkbox)
        checkbox_layout.addWidget(self.plot_mode_checkbox)
        checkbox_layout.addWidget(self.buffer_mode_checkbox)
        checkbox_layout.addStretch()

        # แบ่งครึ่งซ้าย/ขวา: ตาราง | กราฟ
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self.table)
        splitter.addWidget(self.graph)
        splitter.setSizes([400, 700])

        # วางทุกอย่างในหน้าต่าง
        layout.addWidget(self.status_label)
        layout.addWidget(self.source_label)
        layout.addLayout(checkbox_layout)
        layout.addWidget(splitter)
        layout.addLayout(control_layout)
        self.setLayout(layout)

    # ------------------------------
    # เปิด/ปิดการพล็อตกราฟระหว่างทาง
    # ------------------------------
    def toggle_plot_options(self):
        is_enabled = self.enable_plot_checkbox.isChecked()
        self.plot_mode_checkbox.setVisible(is_enabled)
        if not is_enabled:
            self.graph.reset_graph()

    # ------------------------------
    # เลือกไฟล์ปลายทางสำหรับ auto-save/final save
    # ------------------------------
    def select_autosave_file(self):
        if self.monitoring:
            self.status_label.setText("Cannot change auto-save file while monitoring.")
            return
        # ผู้ใช้เลือกได้ทั้ง .xlsx/.csv
        path, _ = QFileDialog.getSaveFileName(
            self, "Select Auto-Save File", "", "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        if path:
            self.auto_save_path = path
            self.auto_save_file_label.setText(os.path.basename(path))
            self.status_label.setText(f"Auto-save file selected: {os.path.basename(path)}")
        else:
            self.auto_save_path = None
            self.auto_save_file_label.setText("No file selected")
            self.status_label.setText("Auto-save file selection cancelled.")

    # ------------------------------
    # ล้างตาราง+กราฟ และสถานะข้อมูลในหน่วยความจำ
    # ------------------------------
    def reset_table(self):
        self.data.clear()
        self.buffered_data.clear()
        self.table.setRowCount(0)
        self.graph.reset_graph()
        self.status_label.setText("Status: Table and graph reset.")
        self.source_label.setText("")

    # ------------------------------
    # ตรวจหาโปรเซสที่จะติดตาม
    # 1) ลองอ่าน PID จาก C:\temp\training_pid.txt (สำหรับ MATLAB)
    # 2) ถ้าไม่พบ: ไล่หาโปรเซส python ที่รัน .py อยู่ (สำหรับสคริปต์เทรน)
    # ------------------------------
    def detect_training_process(self):
        try:
            with open("C:\\temp\\training_pid.txt", "r") as f:
                pid = int(f.read().strip())
                proc = psutil.Process(pid)
                if proc.is_running():
                    cmd = ' '.join(proc.cmdline())
                    self.training_source = f"MATLAB (PID: {pid}) CMD: {cmd}"
                    self.training_pid = pid
                    return True
        except (FileNotFoundError, ValueError, psutil.NoSuchProcess):
            pass

        my_pid = psutil.Process().pid
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if proc.pid == my_pid:
                    continue
                cmdline_args = proc.info.get('cmdline')
                if not cmdline_args:
                    continue
                name = (proc.info['name'] or '').lower()
                # มองหา python + อาร์กิวเมนต์ลงท้าย .py
                if "python" in name and any(str(arg).endswith(".py") for arg in cmdline_args):
                    self.training_source = f"Python: {' '.join(map(str, cmdline_args))}"
                    self.training_pid = proc.pid
                    return True
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                continue
        return False

    # ------------------------------
    # อ่าน CPU/RAM จากโปรเซสเป้าหมาย
    # ------------------------------
    def get_training_process_resource(self, proc):
        try:
            # cpu_percent(interval=None) -> ค่าเฉลี่ยตั้งแต่ครั้งก่อนที่เรียก
            cpu = proc.cpu_percent(interval=None) / psutil.cpu_count()
            ram = proc.memory_info().rss / (1024 * 1024)  # bytes -> MB
            return cpu, ram
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            # โปรเซสหาย/ปิด -> แจ้ง finish 1 ครั้ง
            if not self._finish_emitted:
                self._finish_emitted = True
                self.worker.finish_monitoring_signal.emit("Process terminated.")
            return None, None
        except Exception as e:
            print(f"Error getting process resource: {e}")
            return None, None

    # ------------------------------
    # ดันข้อมูลใน buffer ลงตาราง+กราฟ แล้วเคลียร์ buffer
    # ------------------------------
    def flush_buffer_to_table_and_graph(self):
        if not self.buffered_data:
            return

        # เติมตารางทีละแถว
        for rowdata in self.buffered_data:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(row + 1)))
            for i, val in enumerate(rowdata):
                if i == 0:
                    item_text = self.format_duration(val)  # คอลัมน์เวลา
                else:
                    item_text = f"{val:.2f}" if isinstance(val, float) else str(val)
                self.table.setItem(row, i, QTableWidgetItem(item_text))

        # รวมเข้าชุดข้อมูลหลัก
        self.data.extend(self.buffered_data)

        # ถ้าเปิดพล็อตและไม่ได้เลือก "plot after end" -> วาดแบบเรียลไทม์
        if self.enable_plot_checkbox.isChecked() and not self.plot_mode_checkbox.isChecked():
            is_real_time_mode = self.buffer_mode_checkbox.isChecked()
            timestamps = [d[0] for d in self.data]
            cpu_vals = [d[1] for d in self.data]
            ram_vals = [d[2] for d in self.data]
            self.graph.plot(timestamps, cpu_vals, ram_vals, is_real_time_mode)

        # เคลียร์ buffer แล้วเลื่อนตารางไปท้าย
        self.buffered_data.clear()
        self.table.scrollToBottom()

    # ------------------------------
    # ปรับช่วงเวลาการ flush ตามเวลาที่รัน (ลดภาระ UI)
    # ------------------------------
    def get_dynamic_update_interval(self, elapsed_seconds):
        if elapsed_seconds <= 10: return 10
        if elapsed_seconds <= 20: return 2
        if elapsed_seconds <= 60: return 5
        if elapsed_seconds <= 300: return 10
        if elapsed_seconds <= 900: return 20
        return 30

    # ------------------------------
    # ลูปหลักที่ทำงานใน background thread
    # ------------------------------
    def monitor_loop(self):
        proc_obj = None
        while True:
            # ยังไม่เริ่มมอนิเตอร์ -> ถ้าเลือก auto-start และตรวจพบโปรเซส ให้เริ่ม
            if not self.monitoring:
                if self.auto_start_checkbox.isChecked() and self.detect_training_process():
                    self.start_monitoring()
                    try:
                        proc_obj = psutil.Process(self.training_pid)
                        proc_obj.cpu_percent(interval=None)  # prime CPU counter
                        time.sleep(self.sampling_rate)
                    except psutil.NoSuchProcess:
                        if not self._finish_emitted:
                            self._finish_emitted = True
                            self.worker.finish_monitoring_signal.emit("Process not found.")
                        proc_obj = None
                    except Exception as e:
                        print(f"Error starting monitoring: {e}")
                        if not self._finish_emitted:
                            self._finish_emitted = True
                            self.worker.finish_monitoring_signal.emit(f"Error starting monitoring: {e}")
                else:
                    time.sleep(0.5)
                    continue

            # ถ้าโปรเซสตาย -> แจ้ง finish ครั้งเดียว
            if not psutil.pid_exists(self.training_pid):
                if not self._finish_emitted:
                    self._finish_emitted = True
                    self.worker.finish_monitoring_signal.emit("Process terminated.")
                time.sleep(0.2)
                continue

            start_of_loop = time.time()
            cpu, ram = self.get_training_process_resource(proc_obj)

            if cpu is not None:
                # เวลา ณ session ปัจจุบัน + เวลาสะสมก่อนหน้า -> ทำให้แกน X ต่อเนื่องข้าม autosave/reset
                current_session_elapsed = time.time() - self.training_start_time
                self.buffered_data.append(
                    (self.total_elapsed_time + current_session_elapsed, cpu, ram, self.training_source)
                )

                # ครบ 1 ชั่วโมง -> autosave กลางทาง แล้ว reset session 3600
                if current_session_elapsed >= 3600.0 and len(self.buffered_data) > 0:
                    self.auto_save_data()

                # โหมด flush
                is_real_time_mode = self.buffer_mode_checkbox.isChecked()
                if is_real_time_mode:
                    # flush ทันทีทุกครั้งที่มีข้อมูล (real-time)
                    self.worker.update_ui.emit(self.buffered_data, "flush")
                else:
                    # โหมด buffered: flush ครั้งแรกเมื่อครบ 10 วิ หลังจากนั้นปรับช่วงตามเวลาที่รัน
                    elapsed = time.time() - self.training_start_time
                    if not self.initial_buffer_flushed and elapsed >= 10:
                        self.worker.update_ui.emit(self.buffered_data, "flush")
                        self.last_update_time = time.time()
                        self.initial_buffer_flushed = True
                    elif self.initial_buffer_flushed:
                        self.update_interval = self.get_dynamic_update_interval(elapsed)
                        if time.time() - self.last_update_time >= self.update_interval:
                            self.worker.update_ui.emit(self.buffered_data, "flush")
                            self.last_update_time = time.time()

            # นอนให้ครบตาม sampling_rate
            time_spent = time.time() - start_of_loop
            sleep_time = max(0, self.sampling_rate - time_spent)
            time.sleep(sleep_time)

    # ------------------------------
    # รับประกันว่าไฟล์ CSV จะมีหัวตารางบรรทัดแรกเสมอ
    # - ว่าง/ไม่มีไฟล์ -> เขียนหัว
    # - มีข้อมูลแต่ไม่มีหัว -> แทรกหัวด้านบนโดยใช้ temp file
    # ------------------------------
    def _ensure_csv_header(self, path):
        header = ["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"]
        header_line_norm = ",".join(h.replace(", ", ",").strip() for h in header)

        if not os.path.exists(path) or os.path.getsize(path) == 0:
            with open(path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(header)
            return

        # อ่านบรรทัดแรกเพื่อเทียบว่าเป็นหัวแล้วหรือยัง
        try:
            with open(path, mode='r', encoding='utf-8', newline='') as f:
                first = f.readline().strip()
        except Exception:
            first = ""

        first_norm = first.replace(", ", ",").strip()
        if first_norm == header_line_norm:
            return

        # ถ้าไม่มีหัว -> สร้าง temp แล้วใส่หัว + คัดลอกข้อมูลเดิมทับ
        import tempfile, shutil
        fd, temp_path = tempfile.mkstemp(suffix=".csv")
        os.close(fd)
        try:
            with open(temp_path, mode='w', newline='', encoding='utf-8') as out_f:
                writer = csv.writer(out_f)
                writer.writerow(header)
                with open(path, mode='r', encoding='utf-8', newline='') as in_f:
                    shutil.copyfileobj(in_f, out_f)
            shutil.move(temp_path, path)
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass

    # ------------------------------
    # รับประกันว่าไฟล์ XLSX จะมีหัวตารางแถวแรกเสมอ
    # - ไม่มีไฟล์ -> สร้างใหม่ + เขียนหัว
    # - มีไฟล์:
    #     * แถวแรกว่าง -> เขียนหัว
    #     * แถวแรกไม่ใช่หัว -> แทรกแถวด้านบน แล้วเขียนหัว
    # ------------------------------
    def _ensure_xlsx_header(self, path):
        header = ["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"]

        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            wb.save(path)
            return

        try:
            wb = load_workbook(path)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active

        first_row_vals = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        first_row_empty = all(v is None for v in first_row_vals)

        def norm(x):
            return (x or "").replace(", ", ",").strip()

        header_norm = [norm(h) for h in header]
        first_norm = [norm(v) for v in first_row_vals]

        if first_row_empty:
            for c, val in enumerate(header, start=1):
                ws.cell(row=1, column=c, value=val)
        elif first_norm != header_norm:
            ws.insert_rows(1)
            for c, val in enumerate(header, start=1):
                ws.cell(row=1, column=c, value=val)
        # else: มีหัวถูกต้องแล้ว -> ไม่ทำอะไร

        wb.save(path)

    # ------------------------------
    # Auto-save (กลางทาง)
    # - ถ้าไม่เลือกไฟล์ไว้ -> สร้าง CSV อัตโนมัติใน Downloads ชื่อ Data_YYYYMMDD_HHMMSS.csv
    # - ถ้าเลือก .xlsx/.csv -> append ลงไฟล์นั้น โดยบังคับมีหัวตารางก่อนเสมอ
    # - หลังเขียนเสร็จ -> reset session (ล้างตาราง/เวลาสะสมเฉพาะรอบ)
    # ------------------------------
    def auto_save_data(self):
        try:
            # ถ้ายังไม่ตั้ง path -> สร้าง CSV อัตโนมัติใน Downloads
            if not self.auto_save_path:
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self.auto_save_path = os.path.join(downloads_path, f"Data_{timestamp}.csv")
                # อัปเดต label ชื่อไฟล์ (ให้ผู้ใช้รู้ว่าไฟล์ไปลงชื่ออะไร)
                self.worker.update_ui.emit([], "set_autosave_label:" + os.path.basename(self.auto_save_path))

            path = self.auto_save_path

            # เตรียมข้อมูลที่จะบันทึก (ข้อมูลที่แสดงแล้ว + buffer ค้าง)
            all_data_to_save = self.data + self.buffered_data

            if path.lower().endswith('.xlsx'):
                # บังคับหัวตารางก่อน
                self._ensure_xlsx_header(path)
                # append ลงไฟล์
                wb = load_workbook(path)
                ws = wb.active
                for row_data in all_data_to_save:
                    formatted_row = [self.format_duration(row_data[0])] + list(row_data[1:])
                    ws.append(formatted_row)
                wb.save(path)

            elif path.lower().endswith('.csv'):
                # บังคับหัวตารางก่อน
                self._ensure_csv_header(path)
                # append ลงไฟล์
                with open(path, mode='a', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    for row_data in all_data_to_save:
                        formatted_row = [self.format_duration(row_data[0])] + list(row_data[1:])
                        writer.writerow(formatted_row)

            # แจ้งสถานะ + ตั้งธงว่ามี autosave แล้ว + reset session
            self.status_label.setText(f"Auto-saved data to {os.path.basename(path)} and reset.")
            self._autosave_written = True
            self.reset_data_after_save()

        except Exception as e:
            self.status_label.setText(f"Error during auto-save: {e}")

    # ------------------------------
    # หลัง autosave: เพิ่มเวลาสะสมรวม + ล้างข้อมูล session ปัจจุบัน + เริ่มเวลาใหม่
    # ------------------------------
    def reset_data_after_save(self):
        self.total_elapsed_time += (time.time() - self.training_start_time)
        self.data.clear()
        self.buffered_data.clear()
        self.table.setRowCount(0)
        self.training_start_time = time.time()
        self.last_update_time = self.training_start_time
        self.initial_buffer_flushed = False

    # ------------------------------
    # จัดการสัญญาณจาก worker
    # ------------------------------
    def update_ui(self, new_data, action):
        if action == "flush":
            self.flush_buffer_to_table_and_graph()
        elif action.startswith("set_autosave_label:"):
            label_text = action.split(":", 1)[1]
            self.auto_save_file_label.setText(label_text)

    # ------------------------------
    # จบการมอนิเตอร์ -> Final save
    # - ถ้าเคย autosave มาก่อน: append ต่อไฟล์เดิม (ป้องกันข้อมูลซ้ำ)
    # - ถ้าไม่เคย autosave และไม่มี path: สร้าง FinalData_YYYYMMDD_HHMMSS.xlsx ใหม่
    # ------------------------------
    def finish_monitoring(self, message):
        if self._is_finalizing or self._final_written:
            return
        self._is_finalizing = True

        # ปลดล็อก UI บางส่วน
        self.monitoring = False
        self.sampling_spinbox.setEnabled(True)
        self.btn_select_autosave.setEnabled(True)

        self.status_label.setText(f"Status: {message}. Showing final result...")
        self.source_label.setText(f"Finished monitoring: {self.training_source}")

        path = self.auto_save_path

        # มีข้อมูลค้าง และ (มีไฟล์อยู่แล้วหรือเคยนับเวลาใน session ก่อนหน้า)
        if (self.data or self.buffered_data) and (path is not None or self.total_elapsed_time > 0.0):
            if path is None:
                # ถ้ายังไม่มี path เลย -> ตั้งเป็น XLSX สำหรับ final
                downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                path = os.path.join(downloads_path, f"FinalData_{timestamp}.xlsx")
                self.auto_save_path = path

            all_data_to_save = self.data + self.buffered_data

            try:
                # เคย autosave มาก่อน -> append ต่อไฟล์เดิม (ไม่เขียนหัวซ้ำ)
                if self._autosave_written and os.path.exists(path):
                    if path.lower().endswith('.xlsx'):
                        self._ensure_xlsx_header(path)
                        wb = load_workbook(path)
                        ws = wb.active
                        for row_data in all_data_to_save:
                            formatted_row = [self.format_duration(row_data[0])] + list(row_data[1:])
                            ws.append(formatted_row)
                        # แทรกบรรทัดว่าง + ข้อความ source ไว้ท้ายไฟล์ (เหมือนเวอร์ชันฐาน)
                        ws.append([])
                        ws.append(["", "", "", f"Command/Source: {self.training_source}"])
                        wb.save(path)
                        self.status_label.setText(f"Status: Final data appended to {os.path.basename(path)}")

                    elif path.lower().endswith('.csv'):
                        self._ensure_csv_header(path)
                        with open(path, mode='a', newline='', encoding='utf-8') as file:
                            writer = csv.writer(file)
                            for row_data in all_data_to_save:
                                formatted_row = [self.format_duration(row_data[0])] + list(row_data[1:])
                                writer.writerow(formatted_row)
                            writer.writerow([])
                            writer.writerow(["", "", "", f"Command/Source: {self.training_source}"])
                        self.status_label.setText(f"Status: Final data appended to {os.path.basename(path)}")

                # ไม่เคย autosave -> เขียนใหม่เพื่อหลีกเลี่ยงข้อมูลซ้ำ (เขียนหัวด้วย)
                else:
                    if path.lower().endswith('.xlsx'):
                        wb = Workbook()
                        ws = wb.active
                        ws.append(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
                        for row_data in all_data_to_save:
                            formatted_row = [self.format_duration(row_data[0])] + list(row_data[1:])
                            ws.append(formatted_row)
                        ws.append([])
                        ws.append(["", "", "", f"Command/Source: {self.training_source}"])
                        wb.save(path)
                        self.status_label.setText(f"Status: Final data saved to {os.path.basename(path)}")

                    elif path.lower().endswith('.csv'):
                        with open(path, mode='w', newline='', encoding='utf-8') as file:
                            writer = csv.writer(file)
                            writer.writerow(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
                            for row_data in all_data_to_save:
                                formatted_row = [self.format_duration(row_data[0])] + list(row_data[1:])
                                writer.writerow(formatted_row)
                            writer.writerow([])
                            writer.writerow(["", "", "", f"Command/Source: {self.training_source}"])
                        self.status_label.setText(f"Status: Final data saved to {os.path.basename(path)}")

                self._final_written = True

            except Exception as e:
                self.status_label.setText(f"Error saving final data: {e}")

        # อัปเดตตาราง/กราฟ รอบสุดท้าย (ไม่เขียนไฟล์เพิ่ม)
        self.flush_buffer_to_table_and_graph()

        # ถ้าผู้ใช้เลือก plot-after-end -> วาดกราฟสรุปหลังจบ
        if self.enable_plot_checkbox.isChecked() and self.plot_mode_checkbox.isChecked():
            timestamps = [d[0] for d in self.data]
            cpu_vals = [d[1] for d in self.data]
            ram_vals = [d[2] for d in self.data]
            self.graph.plot(timestamps, cpu_vals, ram_vals, is_real_time=False)

        self._is_finalizing = False

    # ------------------------------
    # เริ่มมอนิเตอร์ใหม่ (รีเซ็ตสถานะรอบใหม่)
    # ------------------------------
    def start_monitoring(self):
        self.sampling_rate = self.sampling_spinbox.value()
        self.monitoring = True
        self.reset_table()
        self.training_start_time = time.time()
        self.last_update_time = self.training_start_time
        self.initial_buffer_flushed = False
        self.idle_start_time = None

        # reset flag สำหรับรอบใหม่
        self.total_elapsed_time = 0.0
        self._finish_emitted = False
        self._is_finalizing = False
        self._final_written = False
        self._autosave_written = False

        # ระหว่างมอนิเตอร์ ไม่อยากให้เผลอไปเปลี่ยน sampling/ไฟล์
        self.sampling_spinbox.setEnabled(False)
        self.btn_select_autosave.setEnabled(False)

        self.status_label.setText("Monitoring...")
        self.source_label.setText(f"Monitoring process: {self.training_source}")

    # ------------------------------
    # Export เฉพาะข้อมูลที่อยู่ใน self.data ปัจจุบันเป็น Excel
    # ------------------------------
    def export_excel(self):
        if not self.data:
            self.status_label.setText("Status: No data to export")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if path:
            wb = Workbook()
            ws = wb.active
            ws.append(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
            for row in self.data:
                formatted_row = [self.format_duration(row[0])] + list(row[1:])
                ws.append(formatted_row)
            ws.append([])
            ws.append(["", "", "", f"Command/Source: {self.training_source}"])
            wb.save(path)
            self.status_label.setText(f"Status: Excel saved to {path}")

    # ------------------------------
    # Export เฉพาะข้อมูลที่อยู่ใน self.data ปัจจุบันเป็น CSV
    # ------------------------------
    def export_csv(self):
        if not self.data:
            self.status_label.setText("Status: No data to export")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save CSV File", "", "CSV Files (*.csv)")
        if path:
            with open(path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(["Time (H:MM:SS.ms)", "CPU (%)", "RAM (MB)", "Source"])
                for row in self.data:
                    formatted_row = [self.format_duration(row[0])] + list(row[1:])
                    writer.writerow(formatted_row)
                writer.writerow([])
                writer.writerow(["", "", "", f"Command/Source: {self.training_source}"])
            self.status_label.setText(f"Status: CSV saved to {path}")

    # ------------------------------
    # บันทึกรูปกราฟปัจจุบันเป็น PNG
    # ------------------------------
    def save_graph(self):
        if not self.enable_plot_checkbox.isChecked():
            self.status_label.setText("Status: Graph plotting is disabled.")
            return
        if not self.data:
            self.status_label.setText("Status: No data to save graph")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save Graph as Image", "", "PNG Files (*.png)")
        if path:
            self.graph.figure.savefig(path, dpi=300, bbox_inches='tight')
            self.status_label.setText(f"Status: Graph saved to {path}")


# ------------------------------
# main entry
# ------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MonitorApp()
    win.show()
    sys.exit(app.exec_())
